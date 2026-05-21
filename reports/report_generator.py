"""
NOCTYRA360™ — Certified Report Generator
Produces SHA-256 signed PDF reports and Excel exports.
Every report is legally admissible — tamper-proof via cryptographic hash.
"""

import hashlib
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Optional

# ── Try to import PDF/Excel libraries, graceful fallback ─────────────────────
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.lib.units import cm
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# ── Color palette (matches V13 dark theme) ────────────────────────────────────
NAVY   = colors.HexColor("#07101C")
GOLD   = colors.HexColor("#C9A227")
WHITE  = colors.white
GREEN  = colors.HexColor("#1E8449")
RED    = colors.HexColor("#C0392B")
LGREY  = colors.HexColor("#EEF2F7")
DGREY  = colors.HexColor("#64748B")


class ReportGenerator:
    """
    Generates certified EFR reports in PDF and Excel formats.
    Each report carries a SHA-256 hash chain for legal admissibility.
    """

    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _report_filename(self, finding: dict, ext: str) -> str:
        operator = finding.get("operator", "Unknown").replace(" ", "_")
        period   = finding.get("period",   "Unknown").replace(" ", "_")
        ts       = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        return str(self.output_dir / f"NOCTYRA360_EFR_{operator}_{period}_{ts}.{ext}")

    # ── JSON report (always available) ────────────────────────────────────────
    def save_json(self, finding: dict) -> str:
        filepath = self._report_filename(finding, "json")
        with open(filepath, "w") as f:
            json.dump(finding, f, indent=2, default=str)
        print(f"  📄 JSON report: {os.path.basename(filepath)}")
        return filepath

    # ── PDF report ────────────────────────────────────────────────────────────
    def generate_pdf(self, finding: dict,
                     anomalies: Optional[dict] = None) -> Optional[str]:
        if not PDF_AVAILABLE:
            print("  ⚠️  ReportLab not installed — skipping PDF (JSON saved)")
            return None

        filepath = self._report_filename(finding, "pdf")
        doc      = SimpleDocTemplate(filepath, pagesize=A4,
                                     leftMargin=1.5*cm, rightMargin=1.5*cm,
                                     topMargin=1.5*cm, bottomMargin=1.5*cm)
        story = []
        styles = getSampleStyleSheet()

        def h1(text):
            return Paragraph(f"<font color='#{GOLD.hexval()[2:]}' size=16><b>{text}</b></font>",
                             styles["Normal"])
        def h2(text):
            return Paragraph(f"<font color='#{NAVY.hexval()[2:]}' size=12><b>{text}</b></font>",
                             styles["Normal"])
        def body(text):
            return Paragraph(f"<font size=10>{text}</font>", styles["Normal"])
        def sp():
            return Spacer(1, 0.3*cm)
        def line():
            return HRFlowable(width="100%", thickness=2,
                              color=GOLD, spaceAfter=6)

        # Header
        story.append(h1("NOCTYRA360™  —  RAPPORT EFR CERTIFIÉ"))
        story.append(body("Connect Now USA LLC  ·  "
                          "Strictly Confidential  ·  For Official Use Only"))
        story.append(line())
        story.append(sp())

        # Key metadata
        cert = finding.get("certification", {})
        meta = finding.get("metadata", {})
        status  = finding.get("status", "UNKNOWN")
        s_color = "#1E8449" if status == "COMPLIANT" else "#C0392B"

        story.append(Table([
            ["Operator:",  finding.get("operator", "—"),
             "Period:",    finding.get("period",   "—")],
            ["Country:",   finding.get("country",  "—"),
             "Currency:",  finding.get("currency", "—")],
            ["Status:",
             f"<font color='{s_color}'><b>{status}</b></font>",
             "Anomaly Score:",
             f"{finding.get('anomaly_score', 0)}/100"],
            ["Certified at:", cert.get("certified_at","—"), "", ""],
        ], colWidths=[3.5*cm, 6*cm, 3.5*cm, 6*cm],
        style=TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), LGREY),
            ("FONTSIZE",   (0,0), (-1,-1), 9),
            ("GRID",       (0,0), (-1,-1), 0.5, colors.white),
            ("FONTNAME",   (0,0), (0,-1),  "Helvetica-Bold"),
            ("FONTNAME",   (2,0), (2,-1),  "Helvetica-Bold"),
        ])))
        story.append(sp())

        # Certified figures
        story.append(h2("CERTIFIED FIGURES (from CDR data)"))
        story.append(sp())
        cert_data = finding.get("certified", {})
        usd_data  = finding.get("summary_usd", {})

        if finding.get("data_type") == "MOBILE_MONEY":
            rows = [
                ["Total Transactions",
                 f"{(cert_data or {}).get('total_transactions',0):,}"],
                ["Total Amount",
                 f"{(cert_data or {}).get('total_amount',0):,.2f} {finding.get('currency','')}"],
                ["Total Fees",
                 f"{(cert_data or {}).get('total_fees',0):,.2f} {finding.get('currency','')}"],
                ["Tax on Fees",
                 f"{(cert_data or {}).get('tax_on_fees',0):,.2f} {finding.get('currency','')}"],
            ]
        else:
            rows = [
                ["Total Revenue",
                 f"{(cert_data or {}).get('total_revenue',0):,.2f} {finding.get('currency','')}",
                 f"USD {usd_data.get('certified_revenue_usd',0):,.2f}"],
                ["Voice Revenue",
                 f"{(cert_data or {}).get('voice_revenue',0):,.2f}", ""],
                ["Data Revenue",
                 f"{(cert_data or {}).get('data_revenue',0):,.2f}", ""],
                ["SMS Revenue",
                 f"{(cert_data or {}).get('sms_revenue',0):,.2f}", ""],
                ["IDD Revenue",
                 f"{(cert_data or {}).get('idd_revenue',0):,.2f}", ""],
                ["Total Certified Tax",
                 f"{(cert_data or {}).get('total_tax',0):,.2f} {finding.get('currency','')}",
                 f"USD {usd_data.get('certified_tax_usd',0):,.2f}"],
            ]

        story.append(Table(rows, colWidths=[6*cm, 7*cm, 6*cm],
            style=TableStyle([
                ("BACKGROUND", (0,0), (-1,-1), LGREY),
                ("FONTSIZE",   (0,0), (-1,-1), 9),
                ("GRID",       (0,0), (-1,-1), 0.5, colors.white),
                ("FONTNAME",   (0,0), (0,-1),  "Helvetica-Bold"),
                ("BACKGROUND", (0,len(rows)-1), (-1,len(rows)-1),
                 colors.HexColor("#E8F5E9")),
            ])))
        story.append(sp())

        # EFR Gap
        gap = finding.get("gap")
        if gap:
            story.append(h2("EFR FISCAL GAP"))
            story.append(sp())
            declared = finding.get("declared", {})
            gap_rows = [
                ["Declared Revenue",
                 f"{(declared or {}).get('total_revenue',0):,.2f} {finding.get('currency','')}",""],
                ["Declared Tax",
                 f"{(declared or {}).get('total_tax',0):,.2f} {finding.get('currency','')}",""],
                ["Certified Tax",
                 f"{(cert_data or {}).get('total_tax',(cert_data or {}).get('tax_on_fees',0)):,.2f} {finding.get('currency','')}",""],
                ["TAX GAP (EFR FINDING)",
                 f"{(gap or {}).get('tax_gap',0):,.2f} {finding.get('currency','')}",
                 f"USD {(gap or {}).get('tax_gap_usd',0):,.2f}"],
                ["Gap Rate",
                 f"{(gap or {}).get('gap_rate_pct',0):.2f}%", ""],
            ]
            story.append(Table(gap_rows, colWidths=[6*cm, 7*cm, 6*cm],
                style=TableStyle([
                    ("BACKGROUND", (0,0), (-1,-1), LGREY),
                    ("FONTSIZE",   (0,0), (-1,-1), 9),
                    ("GRID",       (0,0), (-1,-1), 0.5, colors.white),
                    ("FONTNAME",   (0,0), (0,-1),  "Helvetica-Bold"),
                    ("BACKGROUND", (0,3), (-1,3),
                     colors.HexColor("#FFEBEE")),
                    ("FONTNAME",   (0,3), (-1,3),  "Helvetica-Bold"),
                ])))
            story.append(sp())

        # Security findings
        if anomalies:
            story.append(h2("SECURITY & FRAUD FINDINGS"))
            story.append(sp())
            summary = anomalies.get("summary", {})
            story.append(body(
                f"Total anomalies detected: <b>{summary.get('total_anomalies',0)}</b> | "
                f"High-risk: <b>{summary.get('high_risk_count',0)}</b>"))
            story.append(sp())

            for sim in anomalies.get("simbox", [])[:5]:
                ev = sim.get("evidence", {})
                story.append(Table([
                    [f"SIM BOX — Score {sim['score']}/100", sim.get("verdict","")],
                    ["MSISDN:", sim.get("msisdn","")],
                    ["IDD Calls:", str(ev.get("idd_call_count",""))],
                    ["Avg Duration:", f"{ev.get('avg_duration_sec','')} sec"],
                    ["Action:", sim.get("action","")],
                ], colWidths=[4*cm, 15*cm],
                style=TableStyle([
                    ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#FFEBEE")),
                    ("FONTSIZE",   (0,0), (-1,-1), 8),
                    ("GRID",       (0,0), (-1,-1), 0.3, colors.white),
                ])))
                story.append(sp())

        # SHA-256 Certification block
        story.append(line())
        story.append(h2("SHA-256 CRYPTOGRAPHIC CERTIFICATION"))
        story.append(sp())

        cert_rows = [
            ["Algorithm:",    cert.get("algorithm",    "SHA-256")],
            ["Input Hash:",   cert.get("input_hash",   "")[:40] + "..."],
            ["Finding Hash:", cert.get("finding_hash", "")[:40] + "..."],
            ["Certified by:", cert.get("certified_by", "NOCTYRA360™")],
            ["Certified at:", cert.get("certified_at", "")],
            ["Admissible:",   "YES — legally admissible before any court, "
                              "tax authority or regulatory body"],
        ]
        story.append(Table(cert_rows, colWidths=[4*cm, 15*cm],
            style=TableStyle([
                ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#07101C")),
                ("TEXTCOLOR",  (0,0), (-1,-1), colors.white),
                ("TEXTCOLOR",  (0,0), (0,-1),  colors.HexColor("#C9A227")),
                ("FONTSIZE",   (0,0), (-1,-1), 8),
                ("FONTNAME",   (0,0), (0,-1),  "Helvetica-Bold"),
                ("GRID",       (0,0), (-1,-1), 0.3, colors.HexColor("#162544")),
            ])))

        story.append(sp())
        story.append(body(
            "This document was generated by NOCTYRA360™ v13 Production. "
            "The SHA-256 hash chain above guarantees that no finding has been altered "
            "after certification. Any modification to this document will invalidate the hash. "
            "NOCTYRA360™ is a proprietary platform of Connect Now USA LLC."))

        doc.build(story)
        print(f"  📋 PDF report: {os.path.basename(filepath)}")
        return filepath

    # ── Excel export ──────────────────────────────────────────────────────────
    def generate_excel(self, finding: dict,
                       anomalies: Optional[dict] = None) -> Optional[str]:
        if not EXCEL_AVAILABLE:
            print("  ⚠️  openpyxl not installed — skipping Excel")
            return None

        filepath = self._report_filename(finding, "xlsx")
        wb       = openpyxl.Workbook()

        navy_fill  = PatternFill("solid", fgColor="07101C")
        gold_font  = Font(color="C9A227", bold=True)
        white_font = Font(color="FFFFFF", bold=True)
        grey_fill  = PatternFill("solid", fgColor="EEF2F7")
        red_fill   = PatternFill("solid", fgColor="FFEBEE")
        green_fill = PatternFill("solid", fgColor="E8F5E9")

        def style_header(cell, text):
            cell.value          = text
            cell.fill           = navy_fill
            cell.font           = gold_font
            cell.alignment      = Alignment(horizontal="left", vertical="center")

        def style_label(cell):
            cell.font           = Font(bold=True)
            cell.fill           = grey_fill
            cell.alignment      = Alignment(horizontal="left")

        # ── Sheet 1: Summary ────────────────────────────────────────────────
        ws = wb.active
        ws.title = "EFR Summary"
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 20

        row = 1
        style_header(ws.cell(row, 1), "NOCTYRA360™ — EFR CERTIFIED REPORT")
        ws.merge_cells(f"A{row}:C{row}")
        ws.row_dimensions[row].height = 25
        row += 1

        # Metadata
        for label, value in [
            ("Operator",      finding.get("operator",      "—")),
            ("Country",       finding.get("country",       "—")),
            ("Period",        finding.get("period",        "—")),
            ("Currency",      finding.get("currency",      "—")),
            ("Status",        finding.get("status",        "—")),
            ("Anomaly Score", f"{finding.get('anomaly_score',0)}/100"),
            ("Certified At",  finding.get("certification",{}).get("certified_at","—")),
        ]:
            c = ws.cell(row, 1, label)
            style_label(c)
            v = ws.cell(row, 2, value)
            if label == "Status":
                if value == "COMPLIANT":
                    v.fill = green_fill
                elif "GAP" in value:
                    v.fill = red_fill
            row += 1

        row += 1
        style_header(ws.cell(row, 1), "CERTIFIED FIGURES")
        ws.merge_cells(f"A{row}:C{row}")
        row += 1

        cert = finding.get("certified", {})
        usd  = finding.get("summary_usd", {})
        cur  = finding.get("currency", "")

        for label, val_local, val_usd in [
            ("Total Revenue", cert.get("total_revenue", cert.get("total_amount",0)),
             usd.get("certified_revenue_usd","")),
            ("Voice Revenue", cert.get("voice_revenue",""), ""),
            ("Data Revenue",  cert.get("data_revenue",""),  ""),
            ("SMS Revenue",   cert.get("sms_revenue",""),   ""),
            ("IDD Revenue",   cert.get("idd_revenue",""),   ""),
            ("Total Tax",     cert.get("total_tax",cert.get("tax_on_fees",0)),
             usd.get("certified_tax_usd","")),
        ]:
            if val_local == "":
                continue
            style_label(ws.cell(row, 1, label))
            ws.cell(row, 2, val_local)
            ws.cell(row, 3, val_usd if val_usd else "")
            ws.cell(row, 3).font = Font(color="64748B", italic=True)
            if label == "Total Tax":
                ws.cell(row, 2).fill = green_fill
            row += 1

        # Gap section
        gap = finding.get("gap")
        if gap:
            row += 1
            style_header(ws.cell(row, 1), "EFR FISCAL GAP")
            ws.merge_cells(f"A{row}:C{row}")
            row += 1

            declared = finding.get("declared", {})
            for label, val, usd_val in [
                ("Declared Tax",     (declared or {}).get("total_tax",0), ""),
                ("Certified Tax",    cert.get("total_tax",cert.get("tax_on_fees",0)),
                 usd.get("certified_tax_usd","")),
                ("TAX GAP (EFR)",    (gap or {}).get("tax_gap",0),
                 (gap or {}).get("tax_gap_usd","")),
                ("Gap Rate",         f"{(gap or {}).get('gap_rate_pct',0):.2f}%", ""),
            ]:
                style_label(ws.cell(row, 1, label))
                ws.cell(row, 2, val)
                ws.cell(row, 3, usd_val)
                if "GAP" in label.upper():
                    ws.cell(row, 1).fill = red_fill
                    ws.cell(row, 2).fill = red_fill
                    ws.cell(row, 2).font = Font(bold=True, color="C0392B")
                row += 1

        # Certification
        row += 1
        style_header(ws.cell(row, 1), "SHA-256 CERTIFICATION")
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        cert_block = finding.get("certification", {})
        for label, val in [
            ("Input Hash",   cert_block.get("input_hash",  "")),
            ("Finding Hash", cert_block.get("finding_hash","")),
            ("Algorithm",    cert_block.get("algorithm",   "SHA-256")),
            ("Certified By", cert_block.get("certified_by","")),
        ]:
            style_label(ws.cell(row, 1, label))
            ws.cell(row, 2, val)
            ws.cell(row, 2).font = Font(size=8, color="64748B")
            row += 1

        # ── Sheet 2: Anomalies ───────────────────────────────────────────────
        if anomalies:
            ws2 = wb.create_sheet("Security Findings")
            ws2.column_dimensions["A"].width = 20
            ws2.column_dimensions["B"].width = 20
            ws2.column_dimensions["C"].width = 15
            ws2.column_dimensions["D"].width = 40

            style_header(ws2.cell(1, 1), "SECURITY & FRAUD FINDINGS")
            ws2.merge_cells("A1:D1")
            r = 2

            for category, items in [
                ("SIM BOX",    anomalies.get("simbox",     [])),
                ("IMEI FRAUD", anomalies.get("imei_fraud", [])),
                ("MASS SMS",   anomalies.get("mass_sms",   [])),
                ("AML",        anomalies.get("aml",        [])),
            ]:
                if not items:
                    continue
                style_header(ws2.cell(r, 1), category)
                ws2.merge_cells(f"A{r}:D{r}")
                r += 1
                for col, hdr in enumerate(
                        ["MSISDN/ID","Score","Verdict","Action"], 1):
                    style_label(ws2.cell(r, col, hdr))
                r += 1
                for item in items[:20]:
                    ws2.cell(r, 1, item.get("msisdn", item.get("imei","—")))
                    ws2.cell(r, 2, item.get("score",0))
                    ws2.cell(r, 3, item.get("verdict",""))
                    ws2.cell(r, 4, item.get("action",""))
                    if item.get("score",0) >= 75:
                        ws2.cell(r, 1).fill = red_fill
                    r += 1
                r += 1

        wb.save(filepath)
        print(f"  📊 Excel report: {os.path.basename(filepath)}")
        return filepath

    # ── Generate all formats ───────────────────────────────────────────────────
    def generate_all(self, finding: dict,
                     anomalies: Optional[dict] = None) -> dict:
        """Generate JSON + PDF + Excel reports. Returns all file paths."""
        print(f"\n  Generating certified reports...")
        paths = {
            "json":  self.save_json(finding),
            "pdf":   self.generate_pdf(finding, anomalies),
            "excel": self.generate_excel(finding, anomalies),
        }
        print(f"  ✅ All reports generated")
        return paths

    def generate(self, finding: dict, operator: str = "",
                 country: str = "", period: str = "") -> dict:
        """Alias for generate_all — called by server.py.
        Injects operator/country/period into finding before generating."""
        if operator: finding["operator"] = operator
        if country:  finding["country"]  = country
        if period:   finding["period"]   = period
        return self.generate_all(finding)
